"""
views.py - API视图（完整实现V2.0接口）
根据最新接口文档实现，包含成员A/B/D所有专用接口
"""
import os
import uuid
import json
import hashlib
import threading
import asyncio
import secrets
import logging
from pathlib import Path
from functools import wraps
from django.utils import timezone
from django.db import transaction
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.contrib.auth.hashers import make_password, check_password
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django.db.models import Q
from apps.api.models import SeedURL
from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import (
    User, Template, CrawlTask, PageSnapshot, 
    SeedURL, CrawlerConfig, UserTemplateHistory
)
from .serializers import (
    UserSerializer, TemplateSerializer, TemplateListSerializer,
    CrawlTaskSerializer, PageSnapshotSerializer, SeedURLSerializer
)
from .services.ai_service import get_ollama_service
from .services.snapshot_service import PageSnapshotService

logger = logging.getLogger(__name__)

# ==================== 任务控制信号（用于爬虫停止） ====================
TASK_CONTROL_SIGNALS = {}
TASK_CONTROL_LOCK = threading.Lock()

# 在 views.py 文件开头添加以下内容

from rest_framework import viewsets
from .models import PageSnapshot, SeedURL
from .serializers import PageSnapshotSerializer, SeedURLSerializer


# ==================== ViewSet 定义（供 router 使用） ====================

class PageSnapshotViewSet(viewsets.ModelViewSet):
    """网页快照 ViewSet"""
    queryset = PageSnapshot.objects.all().order_by('-created_at')
    serializer_class = PageSnapshotSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        process_status = self.request.query_params.get('process_status')
        if process_status:
            queryset = queryset.filter(process_status=process_status)
        
        task_type = self.request.query_params.get('task_type')
        if task_type:
            queryset = queryset.filter(task_type=task_type)
        
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(url__icontains=search) | Q(markdown__icontains=search)
            )
        return queryset
    
    def create(self, request, *args, **kwargs):
        """创建快照 - 使用 V2.0 逻辑"""
        url = request.data.get('url')
        markdown = request.data.get('markdown')
        raw_html = request.data.get('raw_html')
        task_type = request.data.get('task_type', 'formal')
        user_prompt = request.data.get('user_prompt', '')
        images = request.data.get('images', [])
        
        if not url or not markdown:
            return Response(
                {'code': 400, 'msg': 'url 和 markdown 为必填字段', 'data': None},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 预览任务限制检查
        if task_type == 'preview':
            preview_count = PageSnapshot.objects.filter(task_type='preview').count()
            if preview_count >= 10:
                return Response(
                    {'code': 400, 'msg': '预览任务最多支持10条数据', 'data': None},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        category = request.data.get('category')
        if not category:
            from .services.snapshot_service import PageSnapshotService
            category = PageSnapshotService.auto_category_from_url(url)
        
        # 保存快照
        from .services.snapshot_service import PageSnapshotService
        result = PageSnapshotService.save_or_update(
            url=url,
            markdown=markdown,
            category=category,
            raw_html=raw_html,
            task_type=task_type,
            user_prompt=user_prompt,
            images=images
        )
        
        if result['obj'] is None:
            return Response(
                {'code': 400, 'msg': '预览任务已达上限', 'data': None},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = self.get_serializer(result['obj'])
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'action': result['action'],
                'snapshot': serializer.data
            }
        })


class SeedURLViewSet(viewsets.ModelViewSet):
    """种子URL ViewSet"""
    queryset = SeedURL.objects.all().order_by('-created_at')
    serializer_class = SeedURLSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        school = self.request.query_params.get('school')
        if school:
            queryset = queryset.filter(school=school)
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        return queryset

# ==================== 公共接口（全员可用） ====================

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response({
            'code': 400, 'msg': '用户名和密码不能为空', 'data': None
        }, status=400)
    
    try:
        user = User.objects.get(username=username)
        
        if check_password(password, user.password):
            # 生成唯一 Token
            token = hashlib.md5(f"{username}{secrets.token_hex(16)}".encode()).hexdigest()
            
            # 保存到用户记录
            user.token = token
            user.save()
            
            return Response({
                'code': 200,
                'msg': 'success',
                'data': {
                    'token': token,
                    'user': UserSerializer(user).data
                }
            })
        else:
            return Response({
                'code': 401, 'msg': '密码错误', 'data': None
            }, status=401)
    except User.DoesNotExist:
        return Response({
            'code': 401, 'msg': '用户不存在', 'data': None
        }, status=401)

def token_required(view_func):
    """Token 验证装饰器 - 从 Authorization 头提取并验证 Token"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        auth_header = request.headers.get('Authorization', '')
        token = None
        
        if auth_header.startswith('Bearer '):
            token = auth_header[7:]
        
        if not token:
            return Response({'code': 401, 'msg': '请先登录', 'data': None}, status=401)
        
        # 从数据库查找 Token 对应的用户
        try:
            user = User.objects.get(token=token)
            request.user = user
            return view_func(request, *args, **kwargs)
        except User.DoesNotExist:
            return Response({'code': 401, 'msg': 'Token 无效，请重新登录', 'data': None}, status=401)
    
    return wrapper

@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    """用户注册"""
    username = request.data.get('username')
    password = request.data.get('password')
    email = request.data.get('email', '')
    
    if not username or not password:
        return Response({
            'code': 400, 'msg': '用户名和密码不能为空', 'data': None
        }, status=400)
    
    if User.objects.filter(username=username).exists():
        return Response({
            'code': 400, 'msg': '用户名已存在', 'data': None
        }, status=400)
    
    user = User.objects.create(
        username=username,
        password=make_password(password),
        email=email
    )
    
    return Response({
        'code': 200,
        'msg': '注册成功',
        'data': {'id': user.id, 'username': user.username}
    })

@api_view(['POST'])
@permission_classes([AllowAny])
def send_email_code(request):
    """
    发送邮箱验证码（联调阶段）
    实际生产环境需要对接邮件服务
    """
    email = request.data.get('email')
    
    if not email:
        return Response({
            'code': 400,
            'msg': '邮箱不能为空',
            'data': None
        }, status=400)
    
    # TODO: 实际发送邮件逻辑
    # 联调阶段可临时返回固定验证码 123456
    # 生产环境需要：
    # 1. 生成6位随机验证码
    # 2. 存储到Redis（有效期5分钟）
    # 3. 调用邮件服务发送
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'message': '验证码已发送（联调测试码：123456）',
            'debug_code': '123456'  # 仅开发环境返回，生产环境删除
        }
    })

@api_view(['POST'])
@token_required
def logout(request):
    """退出登录 - 清除 Token"""
    user = request.user
    user.token = None
    user.save()
    
    return Response({
        'code': 200,
        'msg': '退出成功',
        'data': None
    })

@api_view(['GET'])
def get_dashboard_stats(request):
    """全局统计数据（仪表盘）"""
    from datetime import datetime, timedelta
    
    task_total = CrawlTask.objects.count()
    task_completed = CrawlTask.objects.filter(status='completed').count()
    task_failed = CrawlTask.objects.filter(status='failed').count()
    task_running = CrawlTask.objects.filter(status='running').count()
    
    success_rate = round(task_completed / max(task_total, 1) * 100, 2)
    template_total = Template.objects.count()
    total_pages = PageSnapshot.objects.count()
    
    category_stats = dict(
        PageSnapshot.objects.values('category')
        .annotate(count=Count('id'))
        .values_list('category', 'count')
    )
    
    daily_stats = []
    for i in range(6, -1, -1):
        day = (datetime.now() - timedelta(days=i)).date()
        count = PageSnapshot.objects.filter(created_at__date=day).count()
        daily_stats.append({'date': day.isoformat(), 'count': count})
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'tasks': {
                'total': task_total,
                'running': task_running,
                'completed': task_completed,
                'failed': task_failed,
                'success_rate': success_rate
            },
            'templates': {'total': template_total},
            'pages': {
                'total': total_pages,
                'by_category': category_stats,
                'daily_trend': daily_stats
            }
        }
    })
@api_view(['GET'])
def template_stats(request, pk):
    """获取模板的统计信息"""
    try:
        template = Template.objects.get(pk=pk)
    except Template.DoesNotExist:
        return Response({'code': 404, 'msg': '模板不存在', 'data': None}, status=404)
    
    # 获取该模板的所有任务（排除预览任务）
    tasks = CrawlTask.objects.filter(template=template).exclude(task_type='preview')
    
    # 基础统计
    total_tasks = tasks.count()
    completed_tasks = tasks.filter(status='completed').count()
    failed_tasks = tasks.filter(status='failed').count()
    running_tasks = tasks.filter(status='running').count()
    
    success_rate = round(completed_tasks / max(total_tasks, 1) * 100, 2)
    
    # 数据总量
    total_pages = tasks.aggregate(total=models.Sum('total_pages'))['total'] or 0
    success_pages = tasks.aggregate(total=models.Sum('success_pages'))['total'] or 0
    
    # 近7天趋势
    from datetime import datetime, timedelta
    daily_stats = []
    for i in range(6, -1, -1):
        day = datetime.now().date() - timedelta(days=i)
        count = tasks.filter(created_at__date=day).count()
        daily_stats.append({
            'date': day.isoformat(),
            'count': count
        })
    
    # 近7天数据量趋势
    daily_pages = []
    for i in range(6, -1, -1):
        day = datetime.now().date() - timedelta(days=i)
        total = tasks.filter(created_at__date=day).aggregate(
            total=models.Sum('total_pages')
        )['total'] or 0
        daily_pages.append({
            'date': day.isoformat(),
            'pages': total
        })
    
    # 最近5个任务
    recent_tasks = tasks.order_by('-created_at')[:5].values(
        'task_id', 'task_name', 'status', 'total_pages', 'success_pages', 'created_at'
    )
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'template_id': pk,
            'template_name': template.name,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'failed_tasks': failed_tasks,
            'running_tasks': running_tasks,
            'success_rate': success_rate,
            'total_pages': total_pages,
            'success_pages': success_pages,
            'daily_stats': daily_stats,
            'daily_pages': daily_pages,
            'recent_tasks': list(recent_tasks)
        }
    })

@api_view(['GET'])
def get_crawler_status(request):
    """爬虫状态"""
    seed_total = SeedURL.objects.count()
    seed_success = SeedURL.objects.filter(status='success').count()
    seed_failed = SeedURL.objects.filter(status='failed').count()
    seed_pending = SeedURL.objects.filter(status='pending').count()
    
    task_total = CrawlTask.objects.count()
    task_completed = CrawlTask.objects.filter(status='completed').count()
    task_running = CrawlTask.objects.filter(status='running').count()
    task_failed = CrawlTask.objects.filter(status='failed').count()
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'queue_length': seed_pending,
            'active_workers': task_running,
            'seeds': {
                'total': seed_total,
                'success': seed_success,
                'failed': seed_failed,
                'pending': seed_pending
            },
            'tasks': {
                'total': task_total,
                'running': task_running,
                'completed': task_completed,
                'failed': task_failed
            },
            'success_rate': round(seed_success / max(seed_total, 1) * 100, 2)
        }
    })

@api_view(['GET'])
def get_crawl_status(request, task_id):
    """
    查询单个爬虫任务状态
    GET /api/crawl/status/<task_id>/
    """
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'task_id': str(task.task_id),
                'task_name': task.task_name,
                'task_type': task.task_type,
                'status': task.status,
                'seed_url': task.seed_url,
                'max_depth': task.max_depth,
                'total_pages': task.total_pages,
                'success_pages': task.success_pages,
                'failed_pages': task.failed_pages,
                'error_message': task.error_message,
                'report': task.report,
                'created_at': task.created_at,
                'updated_at': task.updated_at,
                'started_at': task.started_at,
                'completed_at': task.completed_at,
            }
        })
    except CrawlTask.DoesNotExist:
        return Response({
            'code': 404,
            'msg': '任务不存在',
            'data': None
        }, status=404)


@api_view(['GET'])
def list_crawl_tasks(request):
    """
    列出所有爬虫任务
    GET /api/crawl/tasks/
    """
    limit = int(request.query_params.get('limit', 50))
    tasks = CrawlTask.objects.all().order_by('-created_at')[:limit]
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'total': CrawlTask.objects.count(),
            'tasks': [
                {
                    'task_id': str(t.task_id),
                    'task_name': t.task_name,
                    'task_type': t.task_type,
                    'status': t.status,
                    'seed_url': t.seed_url,
                    'max_depth': t.max_depth,
                    'total_pages': t.total_pages,
                    'success_pages': t.success_pages,
                    'failed_pages': t.failed_pages,
                    'error_message': t.error_message,
                    'created_at': t.created_at,
                    'updated_at': t.updated_at,
                }
                for t in tasks
            ]
        }
    })

@api_view(['GET'])
def get_logs(request):
    """日志查询 - 支持按任务ID查询"""
    task_id = request.query_params.get('task_id', None)
    log_file = request.query_params.get('file', 'crawl4ai.log')
    lines = min(int(request.query_params.get('lines', 100)), 1000)
    level = request.query_params.get('level', 'ALL').upper()
    
    log_dir = Path(__file__).resolve().parent.parent.parent / 'logs'
    
    # ✅ 如果有 task_id，读取任务专属日志
    if task_id:
        task_log_dir = log_dir / 'tasks'
        log_path = task_log_dir / f'task_{task_id}.log'
        
        if not log_path.exists():
            return Response({
                'code': 200,
                'msg': 'success',
                'data': {
                    'logs': [],
                    'total': 0,
                    'task_id': task_id,
                    'message': '该任务暂无日志'
                }
            })
        
        try:
            with open(log_path, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                recent_lines = all_lines[-lines:]
            
            if level != 'ALL':
                recent_lines = [line for line in recent_lines if f' {level} ' in line or line.startswith(level)]
            
            return Response({
                'code': 200,
                'msg': 'success',
                'data': {
                    'logs': recent_lines,
                    'total': len(recent_lines),
                    'task_id': task_id,
                    'file': str(log_path)
                }
            })
        except Exception as e:
            return Response({
                'code': 500,
                'msg': f'读取日志失败: {str(e)}',
                'data': {
                    'logs': [],
                    'total': 0,
                    'error': str(e)
                }
            }, status=500)
    
    # 如果没有 task_id，读取全局日志
    log_path = log_dir / log_file
    
    if not log_path.exists():
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'logs': [],
                'total': 0,
                'message': '日志文件不存在'
            }
        })
    
    try:
        with open(log_path, 'r', encoding='utf-8') as f:
            all_lines = f.readlines()
            recent_lines = all_lines[-lines:]
        
        if level != 'ALL':
            recent_lines = [line for line in recent_lines if f' {level} ' in line or line.startswith(level)]
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'logs': recent_lines,
                'total': len(recent_lines),
                'file': str(log_path)
            }
        })
    except Exception as e:
        return Response({
            'code': 500,
            'msg': f'读取日志失败: {str(e)}',
            'data': {
                'logs': [],
                'total': 0,
                'error': str(e)
            }
        }, status=500)

@api_view(['GET'])
def get_log_files(request):
    """获取日志文件列表（包含任务日志）"""
    log_dir = Path(__file__).resolve().parent.parent.parent / 'logs'
    
    if not log_dir.exists():
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {'files': []}
        })
    
    files = []
    
    # 全局日志文件
    for f in log_dir.glob('*.log'):
        files.append({
            'name': f.name,
            'size': f.stat().st_size,
            'modified': f.stat().st_mtime,
            'type': 'global'
        })
    
    # 任务日志文件
    task_log_dir = log_dir / 'tasks'
    if task_log_dir.exists():
        for f in task_log_dir.glob('*.log'):
            files.append({
                'name': f.name,
                'size': f.stat().st_size,
                'modified': f.stat().st_mtime,
                'type': 'task'
            })
    
    files.sort(key=lambda x: x['modified'], reverse=True)
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {'files': files}
    })
    
@api_view(['GET'])
def health_check(request):
    """服务健康检查"""
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'status': 'ok',
            'timestamp': timezone.now().isoformat(),
            'services': {'database': 'ok', 'redis': 'ok', 'minio': 'ok'}
        }
    })


# ==================== 成员A专用接口 ====================

@api_view(['GET'])
def get_crawler_config_from_db(request):
    """获取全局爬虫配置 - 成员A使用"""
    configs = CrawlerConfig.objects.filter(enabled=True)
    config_dict = {}
    for cfg in configs:
        config_dict[cfg.key] = cfg.value
    
    default_config = {
        'concurrency': 5,
        'request_delay': 1.0,
        'max_depth': 2,
        'allowed_domains': [],
        'white_list_patterns': [],
        'enable_dead_check': False,
        'max_concurrent': 5,
        'retry_times': 3,
        'timeout': 30
    }
    
    for key, default_value in default_config.items():
        if key not in config_dict:
            config_dict[key] = default_value
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': config_dict
    })


@api_view(['GET'])
def get_pending_seeds(request):
    """获取待爬种子 - 成员A使用"""
    limit = int(request.query_params.get('limit', 10))
    seeds = SeedURL.objects.filter(status='pending')[:limit]
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'count': seeds.count(),
            'seeds': [
                {
                    'id': s.id,
                    'url': s.url,
                    'school': s.school,
                    'category': s.category,
                    'need_render': s.need_render,
                }
                for s in seeds
            ]
        }
    })


@api_view(['POST'])
def update_seed_status(request):
    """更新种子状态 - 成员A使用"""
    url = request.data.get('url')
    status_value = request.data.get('status')
    
    if not url:
        return Response({'code': 400, 'msg': 'url required', 'data': None}, status=400)
    
    if status_value not in ['pending', 'crawling', 'success', 'failed', 'blocked']:
        return Response({'code': 400, 'msg': 'invalid status', 'data': None}, status=400)
    
    try:
        seed = SeedURL.objects.get(url=url)
        seed.status = status_value
        seed.save()
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {'url': url, 'new_status': status_value}
        })
    except SeedURL.DoesNotExist:
        return Response({'code': 404, 'msg': 'seed not found', 'data': None}, status=404)


@api_view(['POST'])
def upload_image(request):
    """图片上传到MinIO - 成员A使用"""
    try:
        from minio import Minio
        from django.conf import settings
        import io
        
        if 'image' in request.FILES:
            image_data = request.FILES['image'].read()
            filename = request.FILES['image'].name
        elif request.data.get('image_base64'):
            base64_str = request.data.get('image_base64')
            if ',' in base64_str:
                base64_str = base64_str.split(',')[1]
            image_data = base64.b64decode(base64_str)
            filename = request.data.get('filename', 'image.jpg')
        else:
            return Response({'code': 400, 'msg': '请提供图片文件', 'data': None}, status=400)
        
        client = Minio(
            settings.MINIO_ENDPOINT,
            access_key=settings.MINIO_ACCESS_KEY,
            secret_key=settings.MINIO_SECRET_KEY,
            secure=settings.MINIO_SECURE
        )
        
        bucket = 'crawl4ai'
        if not client.bucket_exists(bucket):
            client.make_bucket(bucket)
        
        file_hash = hashlib.md5(image_data).hexdigest()
        ext = filename.split('.')[-1] if '.' in filename else 'jpg'
        object_name = f"images/{file_hash}.{ext}"
        
        client.put_object(
            bucket,
            object_name,
            io.BytesIO(image_data),
            len(image_data),
            content_type=f'image/{ext}'
        )
        
        image_url = f"http://{settings.MINIO_ENDPOINT}/{bucket}/{object_name}"
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'success': True,
                'url': image_url,
                'image_id': file_hash,
                'filename': object_name
            }
        })
        
    except Exception as e:
        logger.error(f"图片上传失败: {str(e)}")
        return Response({'code': 500, 'msg': str(e), 'data': None}, status=500)


@api_view(['POST'])
def save_page_snapshot(request):
    """保存网页快照 - 成员A使用（核心接口）"""
    url = request.data.get('url')
    markdown = request.data.get('markdown')
    task_type = request.data.get('task_type', 'formal')
    user_prompt = request.data.get('user_prompt', '')
    task_id = request.data.get('task_id')
    
    if not url or not markdown:
        return Response({
            'code': 400, 
            'msg': 'url 和 markdown 为必填字段', 
            'data': None
        }, status=400)
    
    category = request.data.get('category') or PageSnapshotService.auto_category_from_url(url)
    
    task = None
    if task_id:
        try:
            task = CrawlTask.objects.get(task_id=task_id)
        except CrawlTask.DoesNotExist:
            pass
    
    result = PageSnapshotService.save_or_update(
        url=url, 
        markdown=markdown, 
        category=category,
        task=task,
        task_type=task_type,
        user_prompt=user_prompt
    )
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'action': result['action'],
            'snapshot_id': result['obj'].id,
            'process_status': result['obj'].process_status
        }
    })


@api_view(['POST'])
def start_crawl_task_api(request):
    """启动爬虫任务 - 成员A使用"""
    seed_url = request.data.get('seed_url')
    max_depth = request.data.get('max_depth', 2)
    config = request.data.get('config', {})
    task_type = request.data.get('task_type', 'formal')
    user_prompt = request.data.get('user_prompt', '')
    ai_model = request.data.get('ai_model', 'qwen2:7b')
    ai_api_url = request.data.get('ai_api_url', 'http://127.0.0.1:11434')
    template_id = request.data.get('template_id', None)  # ✅ 新增
    if not seed_url:
        return Response({'code': 400, 'msg': 'seed_url不能为空', 'data': None}, status=400)
    
    # 预览任务限制最多10条
    if task_type == 'preview':
        existing_count = PageSnapshot.objects.filter(task_type='preview').count()
        if existing_count >= 10:
            return Response({
                'code': 400, 
                'msg': '预览任务最多支持10条数据，请删除旧预览任务后重试', 
                'data': None
            }, status=400)
    # ✅ 获取模板对象
    template_obj = None
    if template_id:
        try:
            template_obj = Template.objects.get(pk=template_id)
        except Template.DoesNotExist:
            return Response({'code': 404, 'msg': '模板不存在', 'data': None}, status=404)
        
    task = CrawlTask.objects.create(
        seed_url=seed_url,
        max_depth=max_depth,
        task_type=task_type,
        template=template_obj,  # ✅ 关联模板
        status='pending'
    )
    
    # 启动后台爬虫线程
    thread = threading.Thread(
        target=_run_async_crawl,
        args=(str(task.task_id), seed_url, max_depth, config),
        daemon=True
    )
    thread.start()
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'task_id': str(task.task_id),
            'status': 'pending',
            'message': '爬虫任务已启动'
        }
    })


@api_view(['POST'])
def report_task_result(request, task_id):
    """上报任务执行结果 - 成员A使用"""
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        
        task.status = request.data.get('status', 'completed')
        task.total_pages = request.data.get('total_pages', 0)
        task.success_pages = request.data.get('success_pages', 0)
        task.failed_pages = request.data.get('failed_pages', 0)
        task.report = request.data.get('report', '')
        if request.data.get('error_message'):
            task.error_message = request.data.get('error_message')
        task.updated_at = timezone.now()
        task.save()
        
        # 更新种子状态
        try:
            seed = SeedURL.objects.get(url=task.seed_url)
            seed.status = 'success' if task.status == 'completed' else 'failed'
            seed.save()
        except SeedURL.DoesNotExist:
            pass
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {'task_id': str(task.task_id), 'status': task.status}
        })
        
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': 'Task not found', 'data': None}, status=404)


# 在 views.py 的 generate_rules 接口中添加保存逻辑

@api_view(['POST'])
def generate_rules(request):
    """P2新增：AI生成采集规则 - 成员A专属"""
    ai_model = request.data.get('ai_model', 'qwen2:7b')
    ai_api_url = request.data.get('ai_api_url', 'http://127.0.0.1:11434')
    user_prompt = request.data.get('user_prompt', '')
    html_skeleton = request.data.get('html_skeleton', '')
    template_id = request.data.get('template_id', None)  # 新增：模板ID
    
    if not user_prompt:
        return Response({
            'code': 400, 'msg': 'user_prompt不能为空', 'data': None
        }, status=400)
    
    ollama = get_ollama_service(api_url=ai_api_url, model=ai_model)
    result = ollama.generate_rules(user_prompt, html_skeleton)
    
    # 如果提供了 template_id，保存规则到模板
    if template_id and result.get('rule_content'):
        try:
            template = Template.objects.get(pk=template_id)
            # 添加新字段到模板模型（需要在模型中添加 crawler_rule 字段）
            template.crawler_rule = result.get('rule_content')
            template.save()
        except Template.DoesNotExist:
            pass
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': result
    })
# 在 views.py 中添加新接口

@api_view(['POST'])
def template_save_rule(request, pk):
    """保存爬虫规则到模板"""
    try:
        template = Template.objects.get(pk=pk)
        crawler_rule = request.data.get('crawler_rule', '')
        
        if crawler_rule:
            template.crawler_rule = crawler_rule
            template.rule_generated_at = timezone.now()
            template.save()
            
            return Response({
                'code': 200,
                'msg': 'success',
                'data': {
                    'template_id': pk,
                    'crawler_rule': crawler_rule[:100] + '...' if len(crawler_rule) > 100 else crawler_rule,
                    'rule_generated_at': template.rule_generated_at
                }
            })
        else:
            return Response({
                'code': 400,
                'msg': '规则内容不能为空',
                'data': None
            }, status=400)
            
    except Template.DoesNotExist:
        return Response({
            'code': 404,
            'msg': '模板不存在',
            'data': None
        }, status=404)
# ==================== 成员B专用接口 ====================

@api_view(['POST'])
def update_clean_status(request):
    """P2新增：上报AI清洗结果 - 成员B使用"""
    snapshot_id = request.data.get('snapshot_id')
    process_status = request.data.get('process_status')
    extracted_data = request.data.get('extracted_data', {})
    error_info = request.data.get('error_info', '')
    
    if not snapshot_id:
        return Response({'code': 400, 'msg': 'snapshot_id不能为空', 'data': None}, status=400)
    
    if process_status not in ['ai_cleaned', 'error']:
        return Response({'code': 400, 'msg': 'process_status必须为ai_cleaned或error', 'data': None}, status=400)
    
    try:
        snapshot = PageSnapshot.objects.get(id=snapshot_id)
        snapshot.process_status = process_status
        snapshot.extracted_data = extracted_data
        if error_info:
            snapshot.error_info = error_info
        snapshot.save()
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'snapshot_id': snapshot_id,
                'process_status': process_status,
                'updated': True
            }
        })
    except PageSnapshot.DoesNotExist:
        return Response({'code': 404, 'msg': '快照不存在', 'data': None}, status=404)


# ==================== 成员D专用接口（前端） ====================

# ---------- 模板管理 ----------

@api_view(['GET'])
def template_list(request):
    """获取模板列表（支持分类筛选）"""
    queryset = Template.objects.all()
    
    category = request.query_params.get('category')
    if category:
        queryset = queryset.filter(category=category)
    
    search = request.query_params.get('search', '')
    if search:
        queryset = queryset.filter(name__icontains=search)
    
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 20))
    
    paginator = Paginator(queryset, page_size)
    results = paginator.get_page(page)
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'count': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'results': TemplateListSerializer(results, many=True).data
        }
    })


@api_view(['POST'])
@token_required  # ✅ 添加这行
def template_create(request):
    """新建模板（含新增字段）"""
    name = request.data.get('name')
    seed_url = request.data.get('seed_url')
    
    if not name or not seed_url:
        return Response({'code': 400, 'msg': 'name和seed_url为必填字段', 'data': None}, status=400)
    
    tags = request.data.get('tags', [])
    if len(tags) > 5:
        return Response({'code': 400, 'msg': '标签最多5个', 'data': None}, status=400)
    for tag in tags:
        if len(tag) > 8:
            return Response({'code': 400, 'msg': f'标签"{tag}"超过8个字', 'data': None}, status=400)
    
    # ✅ 现在 request.user 是真实用户
    template = Template.objects.create(
        name=name,
        seed_url=seed_url,
        tags=tags,
        category=request.data.get('category', 'other'),
        ai_model=request.data.get('ai_model', 'qwen2:7b'),
        ai_api_url=request.data.get('ai_api_url', 'http://127.0.0.1:11434'),
        ai_api_key=request.data.get('ai_api_key', ''),
        user_prompt=request.data.get('user_prompt', ''),
        description=request.data.get('description', ''),
        config=request.data.get('config', {}),
        created_by=request.user  # ✅ 直接使用，因为已经验证了
    )
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': TemplateSerializer(template).data
    }, status=201)

# views.py
@api_view(['GET'])
def template_detail(request, pk):
    """获取模板详情（全量字段）"""
    try:
        template = Template.objects.get(pk=pk)
        serializer = TemplateSerializer(template)
        return Response({
            'code': 200,
            'msg': 'success',
            'data': serializer.data  # ✅ 确保返回 data 字段
        })
    except Template.DoesNotExist:
        return Response({
            'code': 404, 
            'msg': '模板不存在', 
            'data': None
        }, status=404)

@api_view(['PUT'])
@token_required  
def template_update(request, pk):
    """更新模板（含新增字段）"""
    try:
        template = Template.objects.get(pk=pk)
        
        # 更新字段
        template.name = request.data.get('name', template.name)
        template.seed_url = request.data.get('seed_url', template.seed_url)
        template.tags = request.data.get('tags', template.tags)
        template.category = request.data.get('category', template.category)
        template.ai_model = request.data.get('ai_model', template.ai_model)
        template.ai_api_url = request.data.get('ai_api_url', template.ai_api_url)
        template.ai_api_key = request.data.get('ai_api_key', template.ai_api_key)
        template.user_prompt = request.data.get('user_prompt', template.user_prompt)
        template.description = request.data.get('description', template.description)
        template.config = request.data.get('config', template.config)
        
        # ✅ 关键修复：如果状态是 rejected，重新提交时改为 pending
        if template.status == 'rejected':
            template.status = 'pending'
            template.review_comment = ''  # 清空驳回原因
        
        template.save()
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': TemplateSerializer(template).data
        })
    except Template.DoesNotExist:
        return Response({'code': 404, 'msg': '模板不存在', 'data': None}, status=404)

@api_view(['DELETE'])
def template_delete(request, pk):
    """删除模板"""
    try:
        template = Template.objects.get(pk=pk)
        template.delete()
        return Response({'code': 200, 'msg': 'success', 'data': None})
    except Template.DoesNotExist:
        return Response({'code': 404, 'msg': '模板不存在', 'data': None}, status=404)


@api_view(['GET'])
@token_required
def template_history(request):
    """P1新增：个人中心-历史模板"""
    # 装饰器已验证，直接使用 request.user
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 20))
    
    history = UserTemplateHistory.objects.filter(user=request.user).select_related('template')
    paginator = Paginator(history, page_size)
    results = paginator.get_page(page)
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'count': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'results': [
                {
                    'id': h.id,
                    'template_id': h.template.id,
                    'template_name': h.template.name,
                    'template_category': h.template.category,
                    'used_at': h.used_at
                }
                for h in results
            ]
        }
    })


# ---------- 任务控制 ----------

@api_view(['POST'])
def start_task(request):
    """启动采集任务（核心改造，区分预览/正式）"""
    print("=" * 60)
    print("📥 收到 start_task 请求")
    print(f"📥 request.data: {request.data}")
    print("=" * 60)
    
    template_id = request.data.get('template_id')
    original_task_id = request.data.get('original_task_id')  # ✅ 新增：原任务ID
    task_type = request.data.get('task_type', 'formal')
    user_prompt = request.data.get('user_prompt', '')
    ai_model = request.data.get('ai_model', 'qwen2:7b')
    ai_api_url = request.data.get('ai_api_url', 'http://127.0.0.1:11434')
    ai_api_key = request.data.get('ai_api_key', '')
    generated_rule = request.data.get('generated_rule', '')
    
    # ✅ 如果传入了 original_task_id，从原任务获取模板和配置
    if original_task_id and not template_id:
        try:
            original_task = CrawlTask.objects.get(task_id=original_task_id)
            if original_task.template:
                template_id = original_task.template.id
                print(f"✅ 从原任务获取模板ID: {template_id}")
                # 继承原任务的其他配置
                if not user_prompt and original_task.template.user_prompt:
                    user_prompt = original_task.template.user_prompt
                if not generated_rule and original_task.template.crawler_rule:
                    generated_rule = original_task.template.crawler_rule
            else:
                return Response({
                    'code': 400,
                    'msg': '原任务没有关联模板，无法重新执行',
                    'data': None
                }, status=400)
        except CrawlTask.DoesNotExist:
            return Response({
                'code': 404,
                'msg': '原任务不存在',
                'data': None
            }, status=404)
    
    # 预览任务限制最多10条
    if task_type == 'preview':
        existing_count = PageSnapshot.objects.filter(task_type='preview').count()
        if existing_count >= 10:
            return Response({
                'code': 400, 
                'msg': '预览任务最多支持10条数据，请删除旧预览任务后重试', 
                'data': None
            }, status=400)
    
    seed_url = None
    template_name = None
    template_obj = None
    
    if template_id:
        try:
            template_obj = Template.objects.get(pk=template_id)
            seed_url = template_obj.seed_url
            template_name = template_obj.name
            template_obj.usage_count += 1
            template_obj.save()
            
            # 记录历史使用
            if request.user.is_authenticated:
                UserTemplateHistory.objects.get_or_create(
                    user=request.user,
                    template=template_obj
                )
            
            if not user_prompt and template_obj.user_prompt:
                user_prompt = template_obj.user_prompt
            
        except Template.DoesNotExist:
            return Response({'code': 404, 'msg': '模板不存在', 'data': None}, status=404)
    else:
        seed_url = request.data.get('seed_url')
    
    if not seed_url:
        return Response({'code': 400, 'msg': 'seed_url不能为空', 'data': None}, status=400)
    
    from datetime import datetime
    task_name = f"{template_name or '采集任务'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    config = request.data.get('config', {'max_depth': 2, 'max_concurrent': 5})
    
    # ✅ 关键修复：确保 template 对象正确关联
    task = CrawlTask.objects.create(
        task_name=task_name,
        task_type=task_type,
        template=template_obj,  # ← 这里使用 template 对象
        seed_url=seed_url,
        max_depth=config.get('max_depth', 2),
        generated_rule=generated_rule,
        status='pending'
    )
    
    # ✅ 打印日志确认关联
    print(f"✅ 创建任务: {task.task_name}, 关联模板: {task.template.name if task.template else '无'}")
    
    # 启动后台爬虫线程
    thread = threading.Thread(
        target=_run_async_crawl,
        args=(str(task.task_id), seed_url, task.max_depth, config),
        daemon=True
    )
    thread.start()
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'task_id': str(task.task_id),
            'task_name': task_name,
            'status': 'pending',
            'created_at': task.created_at
        }
    })

@api_view(['POST'])
def pause_task(request, task_id):
    """暂停任务"""
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        
        if task.status != 'running':
            return Response({'code': 400, 'msg': '只有运行中的任务可以暂停', 'data': None}, status=400)
        
        with TASK_CONTROL_LOCK:
            if task_id in TASK_CONTROL_SIGNALS:
                TASK_CONTROL_SIGNALS[task_id]['pause_event'] = True
        
        task.status = 'paused'
        task.save()
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {'task_id': task_id, 'status': 'paused'}
        })
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': '任务不存在', 'data': None}, status=404)


@api_view(['POST'])
def stop_task(request, task_id):
    """停止任务"""
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        
        # ✅ 检查状态
        if task.status not in ['running', 'paused', 'pending']:
            return Response({
                'code': 400, 
                'msg': f'只有运行中、等待中或暂停的任务可以停止，当前状态: {task.status}', 
                'data': None
            }, status=400)
        
        # ✅ 设置停止信号
        with TASK_CONTROL_LOCK:
            if task_id in TASK_CONTROL_SIGNALS:
                TASK_CONTROL_SIGNALS[task_id]['stop_event'] = True
                TASK_CONTROL_SIGNALS[task_id]['is_stop'] = True
        
        # ✅ 更新状态
        task.status = 'stopped'
        task.save()
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {'task_id': task_id, 'status': 'stopped'}
        })
    except CrawlTask.DoesNotExist:
        return Response({
            'code': 404, 
            'msg': '任务不存在', 
            'data': None
        }, status=404)
    except Exception as e:
        return Response({
            'code': 500,
            'msg': f'停止失败: {str(e)}',
            'data': None
        }, status=500)

@api_view(['DELETE'])
def delete_task(request, task_id):
    """删除任务"""
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        # 同时删除关联的页面快照
        PageSnapshot.objects.filter(task=task).delete()
        task.delete()
        
        return Response({'code': 200, 'msg': 'success', 'data': None})
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': '任务不存在', 'data': None}, status=404)


# ---------- 任务查询 ----------

@api_view(['GET'])
def task_list_api(request):
    """任务列表（支持状态筛选）"""
    status_filter = request.query_params.get('status', '')
    task_type = request.query_params.get('task_type', '')
    include_preview = request.query_params.get('include_preview', 'false')
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 20))
    template_id = request.query_params.get('template_id', None)  
    queryset = CrawlTask.objects.all()
    # ✅ 按模板ID筛选
    if template_id:
        queryset = queryset.filter(template_id=template_id)
    # 默认排除预览任务
    if include_preview.lower() != 'true':
        queryset = queryset.exclude(task_type='preview')
    
    if status_filter:
        queryset = queryset.filter(status=status_filter)
    if task_type:
        queryset = queryset.filter(task_type=task_type)
    
    paginator = Paginator(queryset, page_size)
    results = paginator.get_page(page)
    
    result_list = []
    for t in results:
        duration = None
        if t.created_at and t.updated_at and t.status in ['completed', 'stopped', 'failed']:
            delta = t.updated_at - t.created_at
            minutes = delta.total_seconds() // 60
            seconds = delta.total_seconds() % 60
            duration = f"{int(minutes):02d}:{int(seconds):02d}"
        elif t.status == 'running' and t.created_at:
            delta = timezone.now() - t.created_at
            minutes = delta.total_seconds() // 60
            seconds = delta.total_seconds() % 60
            duration = f"{int(minutes):02d}:{int(seconds):02d}"
        
        template_obj = t.template
        template_id_val = template_obj.id if template_obj else t.template_id
        
        result_list.append({
            'task_id': str(t.task_id),
            'task_name': t.task_name or f"任务_{t.created_at.strftime('%Y%m%d_%H%M%S')}",
            'task_type': t.task_type,
            'generated_rule': t.generated_rule,
            'template_id': template_id_val,      # ✅ 新增：模板ID
            'template_name': template_obj.name if template_obj else None,
            'status': t.status,
            'duration': duration or '00:00',
            'progress': f"{t.success_pages}/{t.total_pages}" if t.total_pages > 0 else '0/0',
            'progress_percent': int(t.success_pages / max(t.total_pages, 1) * 100),
            'total_pages': t.total_pages,
            'success_pages': t.success_pages,
            'failed_pages': t.failed_pages,
            'created_at': t.created_at,
            'started_at': t.started_at,
            'completed_at': t.completed_at
        })
    
    return Response({
        'code': 200,
        'msg': 'success',
        'data': {
            'count': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'results': result_list
        }
    })

@api_view(['GET'])
def task_detail_api(request, task_id):
    """任务详情"""
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        return Response({
            'code': 200,
            'msg': 'success',
            'data': CrawlTaskSerializer(task).data
        })
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': '任务不存在', 'data': None}, status=404)


@api_view(['GET'])
def task_progress_api(request, task_id):
    """任务进度轮询"""
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        
        total = max(task.total_pages, 1)
        current = task.success_pages
        
        elapsed = None
        if task.created_at:
            delta = timezone.now() - task.created_at
            minutes = delta.total_seconds() // 60
            seconds = delta.total_seconds() % 60
            elapsed = f"{int(minutes):02d}:{int(seconds):02d}"
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'task_id': str(task.task_id),
                'status': task.status,
                'current': current,
                'total': task.total_pages or 100,
                'percent': int(current / total * 100),
                'message': f"已采集{current}页",
                'elapsed_time': elapsed or '00:00'
            }
        })
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': '任务不存在', 'data': None}, status=404)


@api_view(['GET'])
def task_preview_api(request, task_id):
    """采集数据预览（返回结构化数据 + 原始HTML）"""
    limit = int(request.query_params.get('limit', 10))
    
    try:
        import uuid
        try:
            task_uuid = uuid.UUID(task_id)
        except ValueError:
            task_uuid = task_id
        
        try:
            if isinstance(task_uuid, uuid.UUID):
                task = CrawlTask.objects.get(task_id=task_uuid)
            else:
                task = CrawlTask.objects.get(task_id=task_id)
        except CrawlTask.DoesNotExist:
            return Response({
                'code': 404,
                'msg': f'任务不存在: {task_id}',
                'data': None
            }, status=404)
        
        # 查询关联的页面快照
        pages = PageSnapshot.objects.filter(task=task).order_by('-created_at')[:limit]
        
        preview = []
        raw_html = None  # ✅ 新增：存储原始HTML
        markdown_content = None  # ✅ 新增：存储Markdown
        
        for page in pages:
            # ✅ 保存原始HTML（取第一条记录的）
            if not raw_html and page.raw_html:
                raw_html = page.raw_html
            if not markdown_content and page.markdown:
                markdown_content = page.markdown
            
            if page.extracted_data:
                preview.append({
                    'url': page.url,
                    'category': page.category,
                    'extracted_data': page.extracted_data,
                    'created_at': page.created_at.strftime('%Y-%m-%d %H:%M:%S') if page.created_at else None
                })
            else:
                preview.append({
                    'url': page.url,
                    'category': page.category,
                    'markdown_preview': page.markdown[:500] if page.markdown else '',
                    'created_at': page.created_at.strftime('%Y-%m-%d %H:%M:%S') if page.created_at else None
                })
        
        total = PageSnapshot.objects.filter(task=task).count()
        
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {
                'total': total,
                'preview': preview,
                'raw_html': raw_html,  # ✅ 新增：返回原始HTML
                'markdown': markdown_content,  # ✅ 新增：返回Markdown
                'has_raw': bool(raw_html),  # ✅ 新增：标记是否有原始数据
                'has_markdown': bool(markdown_content)  # ✅ 新增：标记是否有Markdown
            }
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"❌ task_preview_api 错误: {error_detail}")
        return Response({
            'code': 500,
            'msg': f'获取预览数据失败: {str(e)}',
            'data': {
                'total': 0,
                'preview': [],
                'raw_html': None,
                'markdown': None,
                'has_raw': False,
                'has_markdown': False
            }
        }, status=500)

@api_view(['GET'])
def task_download_api(request, task_id):
    """结果下载（仅正式任务可导出）"""
    import csv
    from django.http import HttpResponse
    from io import StringIO
    
    try:
        task = CrawlTask.objects.get(task_id=task_id)
        
        # 预览任务禁止下载
        if task.task_type == 'preview':
            return Response({'code': 403, 'msg': '预览任务不支持下载', 'data': None}, status=403)
        
        pages = PageSnapshot.objects.filter(task=task)
        format_type = request.query_params.get('format', 'json')
        
        if format_type == 'csv':
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(['URL', '分类', '提取数据', '创建时间'])
            
            for page in pages:
                writer.writerow([
                    page.url,
                    page.category,
                    json.dumps(page.extracted_data, ensure_ascii=False) if page.extracted_data else '',
                    page.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="task_{task_id}_result.csv"'
            return response
        else:
            data = [
                {
                    'url': page.url,
                    'category': page.category,
                    'extracted_data': page.extracted_data if page.extracted_data else page.markdown,
                    'created_at': page.created_at
                }
                for page in pages
            ]
            return Response({
                'code': 200,
                'msg': 'success',
                'data': data
            })
            
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': '任务不存在', 'data': None}, status=404)


# ==================== 辅助函数 ====================

def _run_async_crawl(task_id, seed_url, max_depth, config):
    """在独立线程中运行异步爬虫"""
    import sys
    import os
    from pathlib import Path
    
    # ============================================================
    # ✅ 关键修复：确保 sandbox 路径在 sys.path 中
    # ============================================================
    current_file = Path(__file__).resolve()
    BACKEND_ROOT = current_file.parent.parent.parent
    PROJECT_ROOT = BACKEND_ROOT.parent.parent
    sandbox_path = PROJECT_ROOT / "sandbox"
    sandbox_str = str(sandbox_path)
    
    # 确保 sandbox 在 sys.path 最前面
    if sandbox_str in sys.path:
        sys.path.remove(sandbox_str)
    sys.path.insert(0, sandbox_str)
    
    print(f"🔍 sandbox路径: {sandbox_str}")
    print(f"🔍 sys.path前3项: {sys.path[:3]}")
    
    # 验证导入
    try:
        import standalone_crawler
        print("✅ standalone_crawler 导入成功")
    except ImportError as e:
        print(f"❌ standalone_crawler 导入失败: {e}")
    
    # ============================================================
    # 设置Django环境
    # ============================================================
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'edu_backend.settings')
    
    import django
    django.setup()
    
    from django.utils import timezone
    from apps.api.models import CrawlTask, SeedURL
    
    # 在爬虫启动前自动创建种子
    seed, created = SeedURL.objects.get_or_create(
        url=seed_url,
        defaults={
            'status': 'pending',
            'need_render': True,
            'school': 'default',  # ✅ 添加 school 字段
            'category': 'other'   # ✅ 添加 category 字段
        }
    )
    if created:
        print(f"✅ 已自动创建种子: {seed_url}")
    
    # ============================================================
    # 创建任务专属日志
    # ============================================================
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"🚀 爬虫线程启动: task_id={task_id}")
    
    task_logger = logging.getLogger(f'crawl_task_{task_id}')
    task_logger.setLevel(logging.DEBUG)
    task_logger.handlers.clear()
    
    log_dir = Path(__file__).resolve().parent.parent.parent / 'logs' / 'tasks'
    log_dir.mkdir(parents=True, exist_ok=True)
    
    log_file = log_dir / f'task_{task_id}.log'
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    task_logger.addHandler(file_handler)
    
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    task_logger.addHandler(console_handler)
    
    task_logger.info(f"🚀 任务 {task_id} 开始执行")
    task_logger.info(f"📌 种子URL: {seed_url}")
    task_logger.info(f"📏 最大深度: {max_depth}")
    task_logger.info(f"⚙️ 配置: {config}")
    
    # ============================================================
    # 注册控制信号
    # ============================================================
    
    with TASK_CONTROL_LOCK:
        TASK_CONTROL_SIGNALS[task_id] = {
            'stop_event': False,
            'pause_event': False,
            'is_stop': False
        }
    
    # ============================================================
    # 执行爬虫
    # ============================================================
    import asyncio
    import threading
    
    try:
        CrawlTask.objects.filter(task_id=task_id).update(
            status='running',
            started_at=timezone.now()
        )
        task_logger.info(f"✅ 任务状态已更新为 running")
        
        # ✅ 导入爬虫模块（路径已修复）
        try:
            from standalone_crawler.crawler import crawl as run_crawl
            task_logger.info("✅ 爬虫模块导入成功")
        except ImportError as e:
            task_logger.error(f"❌ 导入爬虫模块失败: {str(e)}")
            raise Exception(f"导入爬虫模块失败: {str(e)}")
        
        # 创建 API 客户端
        from standalone_crawler.api_client import APIClient
        api_client = APIClient(base_url='http://127.0.0.1:8000')
        
        # 创建事件循环并运行爬虫
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        task_logger.info(f"🔄 开始爬取: {seed_url}")
        
        crawl_task = loop.create_task(run_crawl(
            seed_url=seed_url,
            max_depth=max_depth,
            max_concurrent=config.get('max_concurrent', 5),
            request_delay=config.get('request_delay', 1.0),
            allowed_domains=config.get('allowed_domains', []),
            white_list_patterns=config.get('white_list_patterns', []),
            enable_dead_check=config.get('enable_dead_check', False),
            api_client=api_client,
            task_id=task_id,
        ))
        
        # 定期检查停止信号
        while not crawl_task.done():
            with TASK_CONTROL_LOCK:
                if task_id in TASK_CONTROL_SIGNALS and TASK_CONTROL_SIGNALS[task_id].get('is_stop'):
                    task_logger.warning(f"⏹️ 任务 {task_id} 被用户停止")
                    crawl_task.cancel()
                    try:
                        loop.run_until_complete(crawl_task)
                    except asyncio.CancelledError:
                        pass
                    raise Exception("任务被用户停止")
            
            try:
                loop.run_until_complete(asyncio.sleep(1))
            except RuntimeError:
                break
            except Exception:
                break
        
        if not crawl_task.cancelled():
            stats = crawl_task.result()
            loop.close()
            
            total_pages = stats.total if hasattr(stats, 'total') else 0
            success_pages = stats.success if hasattr(stats, 'success') else 0
            failed_pages = stats.failed if hasattr(stats, 'failed') else 0
            report = stats.report() if hasattr(stats, 'report') else ""
            
            task_logger.info(f"✅ 爬取完成! 总计: {total_pages}, 成功: {success_pages}, 失败: {failed_pages}")
            
            CrawlTask.objects.filter(task_id=task_id).update(
                status='completed',
                total_pages=total_pages,
                success_pages=success_pages,
                failed_pages=failed_pages,
                report=report,
                completed_at=timezone.now(),
                updated_at=timezone.now()
            )
            task_logger.info(f"✅ 任务状态已更新为 completed")
        
    except Exception as e:
        task_logger.error(f"❌ 爬虫异常: {str(e)}")
        task_logger.exception("详细错误堆栈:")
        
        with TASK_CONTROL_LOCK:
            is_stop = task_id in TASK_CONTROL_SIGNALS and TASK_CONTROL_SIGNALS[task_id].get('is_stop', False)
            status = 'stopped' if is_stop else 'failed'
        
        CrawlTask.objects.filter(task_id=task_id).update(
            status=status,
            error_message=str(e),
            completed_at=timezone.now(),
            updated_at=timezone.now()
        )
        task_logger.info(f"✅ 任务状态已更新为 {status}")
        
    finally:
        with TASK_CONTROL_LOCK:
            TASK_CONTROL_SIGNALS.pop(task_id, None)
        task_logger.info(f"🏁 任务 {task_id} 执行结束")

@api_view(['GET'])
def proxy_html(request):
    """获取页面骨架（供 AI 生成规则使用）"""
    import requests
    from bs4 import BeautifulSoup
    
    url = request.query_params.get('url')
    if not url:
        return Response({'code': 400, 'msg': 'url 不能为空'}, status=400)
    
    try:
        resp = requests.get(url, timeout=10)
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # 提取骨架（只保留 class 和 id 属性）
        def simplify(element, depth=0):
            if depth > 3:
                return ''
            if element.name in ['script', 'style', 'meta', 'link']:
                return ''
            result = ''
            if hasattr(element, 'name') and element.name:
                attrs = []
                if element.get('class'):
                    attrs.append(f"class='{' '.join(element.get('class'))}'")
                if element.get('id'):
                    attrs.append(f"id='{element.get('id')}'")
                attrs_str = ' ' + ' '.join(attrs) if attrs else ''
                result += f"{'  ' * depth}<{element.name}{attrs_str}>"
                if element.string and element.string.strip():
                    result += element.string.strip()
                for child in element.children:
                    result += simplify(child, depth + 1)
                result += f"{'  ' * depth}</{element.name}>"
            return result
        
        skeleton = simplify(soup.body) if soup.body else '<div>示例页面</div>'
        return Response({
            'code': 200,
            'msg': 'success',
            'data': {'skeleton': skeleton[:3000]}  # 限制长度
        })
    except Exception as e:
        return Response({
            'code': 500,
            'msg': str(e),
            'data': {'skeleton': '<div>示例页面结构</div>'}
        })
        
@api_view(['POST'])
def review_template(request, pk):
    """
    审核模板（管理员专用）
    
    Body:
        {
            "action": "approve" | "reject",
            "comment": "驳回原因（驳回时必填）"
        }
    """
    # 权限检查：只有管理员可以审核
    if not request.user.is_authenticated:
        return Response({'code': 401, 'msg': '请先登录'}, status=401)
    
    # 检查是否是管理员（is_staff 或自定义角色）
    if not request.user.is_staff:
        return Response({'code': 403, 'msg': '权限不足，仅管理员可审核'}, status=403)
    
    try:
        template = Template.objects.get(pk=pk)
    except Template.DoesNotExist:
        return Response({'code': 404, 'msg': '模板不存在'}, status=404)
    
    # 检查模板状态
    if template.status != 'pending':
        return Response({'code': 400, 'msg': f'模板已处理，当前状态: {template.status}'}, status=400)
    
    action = request.data.get('action')
    comment = request.data.get('comment', '')
    
    if action == 'approve':
        template.status = 'approved'
        template.review_comment = comment or '审核通过'
        template.reviewed_at = timezone.now()
        template.reviewed_by = request.user
        template.is_public = True  # 通过后公开
        template.save()
        return Response({
            'code': 200,
            'msg': '模板已审核通过',
            'data': {
                'template_id': pk,
                'status': 'approved'
            }
        })
    
    elif action == 'reject':
        if not comment:
            return Response({'code': 400, 'msg': '驳回时必须填写原因'}, status=400)
        template.status = 'rejected'
        template.review_comment = comment
        template.reviewed_at = timezone.now()
        template.reviewed_by = request.user
        template.save()
        return Response({
            'code': 200,
            'msg': '模板已驳回',
            'data': {
                'template_id': pk,
                'status': 'rejected',
                'comment': comment
            }
        })
    
    else:
        return Response({'code': 400, 'msg': 'action 必须是 approve 或 reject'}, status=400)
    
    
# ==================== 导出接口 ====================

@api_view(['GET'])
def task_export_api(request, task_id):
    """
    导出任务结果（支持多种格式）
    GET /api/tasks/<task_id>/export/?format=json|csv|md|txt|html|xml|sql|rss
    """
    import csv
    import json
    from django.http import HttpResponse
    from io import StringIO
    from datetime import datetime
    
    try:
        task = CrawlTask.objects.get(task_id=task_id)
    except CrawlTask.DoesNotExist:
        return Response({'code': 404, 'msg': '任务不存在', 'data': None}, status=404)
    
    format_type = request.query_params.get('format', 'json')
    
    # 获取数据
    pages = PageSnapshot.objects.filter(task=task)
    raw_html = pages.first().raw_html if pages.first() else None
    
    # 构建结构化数据
    structured_data = []
    for page in pages:
        if page.extracted_data:
            structured_data.append(page.extracted_data)
        else:
            structured_data.append({
                'url': page.url,
                'category': page.category,
                'content': page.markdown[:500] if page.markdown else '',
                'created_at': page.created_at.strftime('%Y-%m-%d %H:%M:%S') if page.created_at else None
            })
    
    # ===== JSON =====
    if format_type == 'json':
        response_data = {
            'task_id': str(task.task_id),
            'task_name': task.task_name,
            'created_at': task.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'total': len(structured_data),
            'data': structured_data
        }
        if raw_html:
            response_data['raw_html'] = raw_html[:10000]
        
        response = HttpResponse(
            json.dumps(response_data, ensure_ascii=False, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.json"'
        return response
    
    # ===== CSV =====
    elif format_type == 'csv':
        output = StringIO()
        if structured_data:
            fieldnames = list(structured_data[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(structured_data)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.csv"'
        return response
    
    # ===== TXT =====
    elif format_type == 'txt':
        lines = []
        lines.append("=" * 60)
        lines.append(f"采集结果 - {task.task_name or task.task_id}")
        lines.append("=" * 60)
        lines.append(f"任务ID: {task.task_id}")
        lines.append(f"创建时间: {task.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"数据量: {len(structured_data)} 条")
        lines.append("=" * 60)
        lines.append("")
        
        for i, item in enumerate(structured_data, 1):
            lines.append(f"[记录 {i}]")
            for key, value in item.items():
                lines.append(f"  {key}: {value}")
            lines.append("")
            lines.append("-" * 30)
        
        response = HttpResponse("\n".join(lines), content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.txt"'
        return response
    
    # ===== Markdown =====
    elif format_type == 'md':
        lines = []
        lines.append(f"# 采集结果 - {task.task_name or task.task_id}")
        lines.append("")
        lines.append(f"- **任务ID**: {task.task_id}")
        lines.append(f"- **创建时间**: {task.created_at.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"- **数据量**: {len(structured_data)} 条")
        lines.append("")
        
        for i, item in enumerate(structured_data, 1):
            lines.append(f"## 记录 {i}")
            lines.append("")
            for key, value in item.items():
                lines.append(f"- **{key}**: {value}")
            lines.append("")
            lines.append("---")
            lines.append("")
        
        response = HttpResponse("\n".join(lines), content_type='text/markdown')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.md"'
        return response
    
    # ===== HTML =====
    elif format_type == 'html':
        template = """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>采集结果 - {task_name}</title>
    <style>
        body { font-family: Arial, sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; }
        .header { background: #409EFF; color: white; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        table { width: 100%; border-collapse: collapse; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background: #f2f2f2; }
        .raw-data { background: #f5f5f5; padding: 15px; border-radius: 4px; overflow: auto; max-height: 400px; }
        .footer { margin-top: 40px; text-align: center; color: #999; font-size: 12px; }
        pre { white-space: pre-wrap; word-wrap: break-word; }
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 采集结果</h1>
        <p><strong>任务:</strong> {task_name}</p>
        <p><strong>任务ID:</strong> {task_id}</p>
        <p><strong>创建时间:</strong> {created_at}</p>
        <p><strong>数据量:</strong> {total} 条</p>
    </div>
    
    <h2>📋 结构化数据</h2>
    <table>
        <tr>{headers}</tr>
        {rows}
    </table>
    
    <h2>📄 原始HTML</h2>
    <div class="raw-data">
        <pre>{raw_html}</pre>
    </div>
    
    <div class="footer">
        <p>生成时间: {export_time} | Crawl4AI 采集系统</p>
    </div>
</body>
</html>"""
        
        headers = ""
        rows = ""
        if structured_data:
            keys = list(structured_data[0].keys())
            headers = "".join([f"<th>{k}</th>" for k in keys])
            for item in structured_data[:50]:
                row = "".join([f"<td>{str(item.get(k, ''))}</td>" for k in keys])
                rows += f"<tr>{row}</tr>"
        
        raw_html_escaped = (raw_html or '暂无原始数据').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')[:20000]
        
        content = template.format(
            task_name=task.task_name or f"任务_{str(task.task_id)[:8]}",
            task_id=task.task_id,
            created_at=task.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            total=len(structured_data),
            headers=headers,
            rows=rows,
            raw_html=raw_html_escaped,
            export_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        
        response = HttpResponse(content, content_type='text/html')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.html"'
        return response
    
    # ===== XML =====
    elif format_type == 'xml':
        import xml.dom.minidom as minidom
        from xml.etree import ElementTree as ET
        
        root = ET.Element("results")
        root.set("task_id", str(task.task_id))
        root.set("total", str(len(structured_data)))
        
        for item in structured_data:
            record = ET.SubElement(root, "record")
            for key, value in item.items():
                field = ET.SubElement(record, key)
                field.text = str(value)
        
        xml_str = ET.tostring(root, encoding='unicode')
        dom = minidom.parseString(xml_str)
        pretty_xml = dom.toprettyxml(indent="  ")
        
        response = HttpResponse(pretty_xml, content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.xml"'
        return response
    
    # ===== SQL =====
    elif format_type == 'sql':
        lines = []
        lines.append("-- 采集数据导入")
        lines.append(f"-- 任务ID: {task.task_id}")
        lines.append(f"-- 数据量: {len(structured_data)} 条")
        lines.append("")
        
        if structured_data:
            table_name = f"crawl_data_{str(task.task_id)[:8]}"
            fields = list(structured_data[0].keys())
            lines.append(f"CREATE TABLE IF NOT EXISTS {table_name} (")
            lines.append("    id INTEGER PRIMARY KEY AUTOINCREMENT,")
            for field in fields:
                lines.append(f"    {field} TEXT,")
            lines.append("    created_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            lines.append(");")
            lines.append("")
            
            for item in structured_data:
                values = []
                for field in fields:
                    val = str(item.get(field, '')).replace("'", "''")
                    values.append(f"'{val}'")
                lines.append(f"INSERT INTO {table_name} ({', '.join(fields)}) VALUES ({', '.join(values)});")
        
        response = HttpResponse("\n".join(lines), content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.sql"'
        return response
    
    # ===== RSS =====
    elif format_type == 'rss':
        rss_template = """<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0">
    <channel>
        <title>采集结果 - {task_name}</title>
        <link>http://localhost</link>
        <description>采集数据订阅</description>
        <pubDate>{pub_date}</pubDate>
        {items}
    </channel>
</rss>"""
        
        items = []
        for item in structured_data[:20]:
            title = item.get('title', '未命名') or item.get('name', '未命名')
            desc = str(item.get('content', '') or item.get('description', ''))[:200]
            items.append(f"""
        <item>
            <title>{title}</title>
            <link>{item.get('url', '')}</link>
            <description>{desc}</description>
            <pubDate>{datetime.now().strftime('%a, %d %b %Y %H:%M:%S GMT')}</pubDate>
        </item>
            """)
        
        content = rss_template.format(
            task_name=task.task_name or f"任务_{str(task.task_id)[:8]}",
            pub_date=datetime.now().strftime('%a, %d %b %Y %H:%M:%S GMT'),
            items="".join(items)
        )
        
        response = HttpResponse(content, content_type='application/rss+xml')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.xml"'
        return response
    
    # ===== Excel (XLSX) =====
    elif format_type == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({
                'code': 400,
                'msg': 'Excel导出需要安装 openpyxl: pip install openpyxl',
                'data': None
            }, status=400)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "采集数据"
        
        if structured_data:
            headers = list(structured_data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            for row, item in enumerate(structured_data, 2):
                for col, key in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=str(item.get(key, '')))
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="task_{task_id}.xlsx"'
        return response
    
    else:
        return Response({
            'code': 400,
            'msg': f'不支持的格式: {format_type}。支持: json, csv, xlsx, md, txt, html, xml, sql, rss',
            'data': None
        }, status=400)